#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Dify 模型识别 Agent 双层端到端回归测试脚本。

测试两件事：
1. match 识别情况是否符合预期：
   从 Dify 最终输出中提取每个查询头部的
   model-name / attribute-tag / framwork / hardware，并和 xlsx 的「预期match情况」比较。

2. 最终检索输出是否符合预期：
   用 xlsx 的「预期输出」和 Dify 实际回答比较，支持 <matches> / <shown> 占位符。

依赖：
    pip install requests openpyxl urllib3

推荐用法：
    export DIFY_API_KEY="你的 Dify App API Key"
    python run_dify_e2e_match_regression_v5.py \
      --cases 模型识别Agent_端到端双层回归测试用例_v5.xlsx \
      --output 模型识别Agent_端到端双层回归测试结果_v5.xlsx \
      --sleep 3 \
      --no-verify
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import traceback
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import urllib3
except Exception:  # pragma: no cover
    urllib3 = None


DEFAULT_URL = "https://api.dify.ai/v1/chat-messages"
DEFAULT_CASES = "模型识别Agent_端到端双层回归测试用例_v5.xlsx"
DEFAULT_OUTPUT = "模型识别Agent_端到端双层回归测试结果_v5.xlsx"


@dataclass
class Case:
    row_no: int
    query: str
    expected_match: str
    expected_output: str


def normalize_text(text: Any, *, unify_query_label: bool = False) -> str:
    """归一化文本，用于宽松比较。"""
    if text is None:
        return ""

    s = str(text)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("：", ":").replace("；", ";").replace("，", ",")
    s = s.replace("→", "->")
    s = s.replace("framework:", "framwork:")
    s = s.replace("Framework:", "framwork:")
    s = s.replace("FRAMEWORK:", "framwork:")
    s = s.lower()

    # 统一中文/英文官方、三方标记的常见差异
    s = s.replace("【官方】", "[official]").replace("【三方】", "[third-party]")
    s = s.replace("[官方]", "[official]").replace("[三方]", "[third-party]")
    s = s.replace("third party", "third-party")

    # 可选：只在最终输出比较时统一 query label；match 比较默认不统一，避免中文/英文格式错过。
    if unify_query_label:
        s = re.sub(r"【\s*查询\s*(\d+)\s*】", r"[query \1]", s)
        s = re.sub(r"\[\s*查询\s*(\d+)\s*\]", r"[query \1]", s)
    else:
        s = re.sub(r"【\s*查询\s*(\d+)\s*】", r"【查询\1】", s)
        s = re.sub(r"\[\s*查询\s*(\d+)\s*\]", r"[查询\1]", s)

    s = re.sub(r"\[\s*query\s*(\d+)\s*\]", r"[query \1]", s)
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r" *\n+ *", "\n", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def canonicalize_match(text: str) -> str:
    """
    对 match 情况做字段级归一化：
    - 保留中文/英文 query label 差异，避免语言格式错误被误判通过。
    - 兼容 framework/framwork 拼写。
    - 标准化字段顺序为 model-name / attribute-tag / framwork / hardware。
    """
    lines = []
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        # 去掉后面的命中信息
        line = re.split(r"\s*(?:->|→)\s*", line, maxsplit=1)[0].strip()
        line = line.replace("framework:", "framwork:")
        label = ""
        m_label = re.match(r"(【\s*查询\s*\d+\s*】|\[\s*查询\s*\d+\s*\]|\[\s*query\s*\d+\s*\])", line, re.I)
        if m_label:
            label = normalize_text(m_label.group(1), unify_query_label=False)
            body = line[m_label.end():]
        else:
            body = line

        fields = extract_fields_from_header(body)
        sep = "" if label.startswith("【") else " "
        lines.append(
            f"{label}{sep}model-name: {fields.get('model-name','')} "
            f"attribute-tag: {fields.get('attribute-tag','')} "
            f"framwork: {fields.get('framwork','')} "
            f"hardware: {fields.get('hardware','')}"
        )

    return normalize_text("\n".join(lines), unify_query_label=False)


def extract_fields_from_header(header: str) -> Dict[str, str]:
    """从单行 header 中抽取字段。支持字段为空。"""
    s = str(header or "")
    s = re.sub(r"framework\s*:", "framwork:", s, flags=re.I)
    keys = ["model-name", "attribute-tag", "framwork", "hardware"]

    positions: List[Tuple[str, int, int]] = []
    for m in re.finditer(r"(model-name|attribute-tag|framwork|hardware)\s*:", s, re.I):
        positions.append((m.group(1).lower(), m.start(), m.end()))

    fields = {k: "" for k in keys}
    for idx, (key, _start, end) in enumerate(positions):
        next_start = positions[idx + 1][1] if idx + 1 < len(positions) else len(s)
        value = s[end:next_start].strip()
        fields[key] = normalize_text(value, unify_query_label=False)

    return fields


def extract_match_from_answer(answer: str) -> str:
    """
    从 Dify 最终回答中提取 match 情况。

    目标提取示例：
    【查询1】model-name: qwen3 attribute-tag: coder, 32b framwork:  hardware:
    [query 1] model-name: qwen3.5 attribute-tag: 27b framwork:  hardware:
    """
    if not answer:
        return ""

    result_lines: List[str] = []
    for raw_line in str(answer).splitlines():
        line = raw_line.strip()
        if not line:
            continue

        # 必须包含查询编号和 model-name 才认为是识别头部
        has_query_label = bool(re.search(r"(【\s*查询\s*\d+\s*】|\[\s*查询\s*\d+\s*\]|\[\s*query\s*\d+\s*\])", line, re.I))
        if not has_query_label or "model-name" not in line.lower():
            continue

        # 截掉命中数和后续内容
        header = re.split(r"\s*(?:→|->)\s*", line, maxsplit=1)[0].strip()
        result_lines.append(header)

    return "\n".join(result_lines)


def expected_to_regex(expected: str, *, unify_query_label: bool = False) -> str:
    """把预期输出转换为正则，支持 <matches>/<shown> 通配。"""
    s = normalize_text(expected, unify_query_label=unify_query_label)
    escaped = re.escape(s)
    escaped = escaped.replace(re.escape("<matches>"), r"\d+")
    escaped = escaped.replace(re.escape("<shown>"), r"\d+")
    # 空白宽松
    escaped = escaped.replace(r"\ ", r"\s+")
    escaped = escaped.replace(r"\n", r"\s*")
    return escaped


def compare_output(expected: str, actual: str, mode: str) -> bool:
    if mode == "none":
        return True

    exp = expected or ""
    act = actual or ""

    if mode == "exact":
        return normalize_text(exp, unify_query_label=False) == normalize_text(act, unify_query_label=False)

    if mode == "contains":
        return normalize_text(exp, unify_query_label=True) in normalize_text(act, unify_query_label=True)

    if mode == "placeholder":
        pattern = expected_to_regex(exp, unify_query_label=True)
        actual_norm = normalize_text(act, unify_query_label=True)
        return re.search(pattern, actual_norm, flags=re.I | re.S) is not None

    raise ValueError(f"Unsupported output compare mode: {mode}")


def compare_match(expected_match: str, actual_match: str, mode: str) -> bool:
    if mode == "none":
        return True

    exp = canonicalize_match(expected_match)
    act = canonicalize_match(actual_match)

    if mode == "exact":
        return exp == act

    if mode == "contains":
        return exp in act

    raise ValueError(f"Unsupported match compare mode: {mode}")


def load_cases(path: str, sheet_name: Optional[str] = None) -> List[Case]:
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    required = ["输入", "预期match情况", "预期输出"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError(f"测试用例缺少列: {missing}。当前表头: {headers}")

    idx = {h: headers.index(h) + 1 for h in required}
    cases: List[Case] = []
    for r in range(2, ws.max_row + 1):
        query = ws.cell(r, idx["输入"]).value
        if query is None or str(query).strip() == "":
            continue
        cases.append(
            Case(
                row_no=r,
                query=str(query),
                expected_match=str(ws.cell(r, idx["预期match情况"]).value or ""),
                expected_output=str(ws.cell(r, idx["预期输出"]).value or ""),
            )
        )
    return cases


def extract_answer_from_dify_payload(payload: Dict[str, Any]) -> str:
    """兼容 Dify chat-messages 常见返回结构。"""
    for key in ("answer", "text", "result"):
        if isinstance(payload.get(key), str):
            return payload[key]

    data = payload.get("data")
    if isinstance(data, dict):
        for key in ("answer", "text", "result", "output"):
            if isinstance(data.get(key), str):
                return data[key]

    outputs = payload.get("outputs")
    if isinstance(outputs, dict):
        for key in ("answer", "text", "result", "output"):
            if isinstance(outputs.get(key), str):
                return outputs[key]

    # 兜底：返回 JSON，避免信息丢失
    return json.dumps(payload, ensure_ascii=False, indent=2)


def run_dify(
    query: str,
    *,
    url: str,
    api_key: str,
    user: str,
    timeout: int,
    verify: bool,
    inputs_json: Optional[str],
) -> Dict[str, Any]:
    inputs: Dict[str, Any] = {}
    if inputs_json:
        inputs = json.loads(inputs_json)

    payload = {
        "inputs": inputs,
        "query": query,
        "response_mode": "blocking",
        "conversation_id": "",
        "user": user,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    resp = requests.post(url, headers=headers, json=payload, timeout=timeout, verify=verify)
    if resp.status_code != 200:
        raise RuntimeError(f"HTTP {resp.status_code}: {resp.text}")
    return resp.json()


def write_results(path: str, rows: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "回归测试结果"

    headers = [
        "序号",
        "输入",
        "预期match情况",
        "实际match情况",
        "是否预期match",
        "预期输出",
        "实际输出",
        "是否符合预期",
        "错误信息",
    ]
    ws.append(headers)

    for i, row in enumerate(rows, start=1):
        ws.append([
            i,
            row.get("query", ""),
            row.get("expected_match", ""),
            row.get("actual_match", ""),
            row.get("match_pass", ""),
            row.get("expected_output", ""),
            row.get("actual_output", ""),
            row.get("output_pass", ""),
            row.get("error", ""),
        ])

    # 样式
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        1: 8,
        2: 36,
        3: 72,
        4: 72,
        5: 16,
        6: 88,
        7: 88,
        8: 16,
        9: 42,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 汇总 Sheet
    sm = wb.create_sheet("汇总")
    sm_rows = [
        ["指标", "值"],
        ["总用例数", summary.get("total", 0)],
        ["match 通过数", summary.get("match_pass", 0)],
        ["match 失败数", summary.get("match_fail", 0)],
        ["最终输出通过数", summary.get("output_pass", 0)],
        ["最终输出失败数", summary.get("output_fail", 0)],
        ["双层均通过数", summary.get("both_pass", 0)],
        ["双层均通过率", summary.get("both_pass_rate", "")],
        ["输出文件", path],
    ]
    for r in sm_rows:
        sm.append(r)
    for cell in sm[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    sm.column_dimensions["A"].width = 24
    sm.column_dimensions["B"].width = 50

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Dify 模型识别 Agent 双层端到端回归测试")
    parser.add_argument("--cases", default=DEFAULT_CASES, help="测试用例 xlsx 路径")
    parser.add_argument("--sheet", default=None, help="测试用例 Sheet 名；默认读取第一个 Sheet")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="输出结果 xlsx 路径")
    parser.add_argument("--url", default=DEFAULT_URL, help="Dify chat-messages API URL")
    parser.add_argument("--api-key", default=os.getenv("DIFY_API_KEY", ""), help="Dify App API Key；也可用环境变量 DIFY_API_KEY")
    parser.add_argument("--user", default="python-regression-client", help="Dify user 字段")
    parser.add_argument("--timeout", type=int, default=120, help="单次请求超时时间，秒")
    parser.add_argument("--sleep", type=float, default=2.0, help="每条用例之间的请求间隔，秒")
    parser.add_argument("--retries", type=int, default=2, help="失败重试次数")
    parser.add_argument("--retry-sleep", type=float, default=5.0, help="失败重试间隔，秒")
    parser.add_argument("--limit", type=int, default=0, help="只跑前 N 条，0 表示全量")
    parser.add_argument("--offset", type=int, default=0, help="从第 N 条开始跑，便于断点调试")
    parser.add_argument("--no-verify", action="store_true", help="requests 关闭 SSL verify")
    parser.add_argument("--inputs-json", default=None, help='Dify inputs JSON 字符串，例如 \'{"foo":"bar"}\'')
    parser.add_argument("--match-compare-mode", choices=["exact", "contains", "none"], default="exact", help="match 情况比较方式")
    parser.add_argument("--output-compare-mode", choices=["placeholder", "contains", "exact", "none"], default="placeholder", help="最终输出比较方式")
    args = parser.parse_args()

    if args.no_verify and urllib3 is not None:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    if not args.api_key:
        print("ERROR: 缺少 Dify API Key。请设置环境变量 DIFY_API_KEY 或传入 --api-key。", file=sys.stderr)
        return 2

    cases = load_cases(args.cases, args.sheet)
    if args.offset:
        cases = cases[args.offset:]
    if args.limit and args.limit > 0:
        cases = cases[:args.limit]

    print(f"Loaded {len(cases)} cases from {args.cases}")

    result_rows: List[Dict[str, Any]] = []
    for idx, case in enumerate(cases, start=1):
        print(f"[{idx}/{len(cases)}] query={case.query!r}")

        actual_answer = ""
        actual_match = ""
        error = ""
        for attempt in range(args.retries + 1):
            try:
                payload = run_dify(
                    case.query,
                    url=args.url,
                    api_key=args.api_key,
                    user=args.user,
                    timeout=args.timeout,
                    verify=not args.no_verify,
                    inputs_json=args.inputs_json,
                )
                actual_answer = extract_answer_from_dify_payload(payload)
                actual_match = extract_match_from_answer(actual_answer)
                break
            except Exception as exc:
                error = f"{type(exc).__name__}: {exc}"
                if attempt < args.retries:
                    print(f"  request failed, retry {attempt + 1}/{args.retries}: {error}")
                    time.sleep(args.retry_sleep)
                else:
                    print(f"  request failed finally: {error}")

        try:
            match_ok = compare_match(case.expected_match, actual_match, args.match_compare_mode)
        except Exception as exc:
            match_ok = False
            error = (error + "\n" if error else "") + f"match compare error: {type(exc).__name__}: {exc}"

        try:
            output_ok = compare_output(case.expected_output, actual_answer, args.output_compare_mode)
        except Exception as exc:
            output_ok = False
            error = (error + "\n" if error else "") + f"output compare error: {type(exc).__name__}: {exc}"

        result_rows.append({
            "query": case.query,
            "expected_match": case.expected_match,
            "actual_match": actual_match,
            "match_pass": "是" if match_ok else "否",
            "expected_output": case.expected_output,
            "actual_output": actual_answer,
            "output_pass": "是" if output_ok else "否",
            "error": error,
        })

        if args.sleep > 0 and idx < len(cases):
            time.sleep(args.sleep)

    total = len(result_rows)
    match_pass = sum(1 for r in result_rows if r["match_pass"] == "是")
    output_pass = sum(1 for r in result_rows if r["output_pass"] == "是")
    both_pass = sum(1 for r in result_rows if r["match_pass"] == "是" and r["output_pass"] == "是")

    summary = {
        "total": total,
        "match_pass": match_pass,
        "match_fail": total - match_pass,
        "output_pass": output_pass,
        "output_fail": total - output_pass,
        "both_pass": both_pass,
        "both_pass_rate": f"{both_pass / total:.2%}" if total else "0.00%",
    }

    write_results(args.output, result_rows, summary)

    print("\n==== Summary ====")
    for k, v in summary.items():
        print(f"{k}: {v}")
    print(f"Saved: {args.output}")

    failed = [r for r in result_rows if r["match_pass"] != "是" or r["output_pass"] != "是"]
    if failed:
        print("\nFirst failed cases:")
        for r in failed[:5]:
            print("- query:", r["query"])
            print("  match_pass:", r["match_pass"], "output_pass:", r["output_pass"])
            if r["error"]:
                print("  error:", r["error"])

    return 0 if both_pass == total else 1


if __name__ == "__main__":
    raise SystemExit(main())
