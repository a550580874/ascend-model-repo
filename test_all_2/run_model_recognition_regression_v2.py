#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
模型识别 Agent 回归测试脚本 v2

核心口径：
- 模型系列版本属于 model-name：
  qwen3-30B -> model-name: qwen3; attribute-tag: 30b
  glm-4.7 -> model-name: glm4.7
  deepseek-r1-distill-qwen-32b -> model-name: deepseek-r1; attribute-tag: distill, qwen, 32b
- attribute-tag 只放参数规模、量化、精度、任务/模态、distill 等标签。
"""

import argparse
import importlib
import re
import sys
import traceback
from pathlib import Path
from typing import Any, Callable, Dict, List, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "缺少依赖 openpyxl，请先执行：pip install openpyxl"
    ) from exc


DEFAULT_CASES = "模型识别Agent_全量回归测试用例_v2.xlsx"
DEFAULT_OUTPUT = "模型识别Agent_回归测试结果_v2.xlsx"


def load_function(module_name: str, function_name: str) -> Callable[[str], Any]:
    module = importlib.import_module(module_name)
    func = getattr(module, function_name, None)
    if func is None or not callable(func):
        raise AttributeError(f"模块 {module_name!r} 中找不到可调用函数 {function_name!r}")
    return func


def cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_cases(path: str, sheet_name: str = None) -> List[Tuple[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header = [cell_to_str(c.value) for c in ws[1]]
    try:
        input_idx = header.index("输入") + 1
        expected_idx = header.index("预期输出") + 1
    except ValueError as exc:
        raise ValueError("测试用例表头必须包含两列：输入、预期输出") from exc

    cases: List[Tuple[str, str]] = []
    for row in range(2, ws.max_row + 1):
        query = cell_to_str(ws.cell(row=row, column=input_idx).value)
        expected = cell_to_str(ws.cell(row=row, column=expected_idx).value)
        if not query and not expected:
            continue
        cases.append((query, expected))
    return cases


def normalize_basic(text: Any) -> str:
    """
    基础比较：
    - 转字符串
    - 转小写
    - 去掉首尾空白
    - 把连续空白压缩成一个空格
    """
    s = cell_to_str(text).lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_loose(text: Any) -> str:
    """
    宽松比较：
    - 兼容 framework / framwork 拼写差异
    - 兼容中文字段名
    - 兼容中文/英文标点
    - 兼容 [查询1] / [query 1]
    - 兼容多余空格

    注意：宽松比较不会把 qwen3 拆成 qwen + attribute-tag: 3，也不会把二者视为等价。
    因此它仍然可以校验本次新口径。
    """
    s = normalize_basic(text)

    # 标点统一
    replacements = {
        "：": ":",
        "；": ";",
        "，": ",",
        "（": "(",
        "）": ")",
        "【": "[",
        "】": "]",
    }
    for old, new in replacements.items():
        s = s.replace(old, new)

    # 字段名统一
    field_aliases = {
        "framework:": "framwork:",
        "模型名:": "model-name:",
        "模型名称:": "model-name:",
        "model name:": "model-name:",
        "model_name:": "model-name:",
        "属性标签:": "attribute-tag:",
        "属性:": "attribute-tag:",
        "attribute_tag:": "attribute-tag:",
        "attribute tag:": "attribute-tag:",
        "框架:": "framwork:",
        "适配框架:": "framwork:",
        "硬件:": "hardware:",
        "适配硬件:": "hardware:",
    }
    for old, new in field_aliases.items():
        s = s.replace(old, new)

    # 查询编号统一：中文 [查询1] -> 英文 [query 1]
    s = re.sub(r"\[查询\s*(\d+)\]", r"[query \1]", s)
    s = re.sub(r"\[query(\d+)\]", r"[query \1]", s)
    s = re.sub(r"\[query\s+(\d+)\]", r"[query \1]", s)

    # 分隔符和冒号周围空白统一
    s = re.sub(r"\s*:\s*", ": ", s)
    s = re.sub(r"\s*,\s*", ", ", s)
    s = re.sub(r"\s*;\s*", "; ", s)
    s = re.sub(r"\s+", " ", s).strip()

    return s


def compare_output(expected: str, actual: str, loose: bool = False) -> bool:
    if loose:
        return normalize_loose(expected) == normalize_loose(actual)
    return normalize_basic(expected) == normalize_basic(actual)


def run_cases(
    cases: List[Tuple[str, str]],
    func: Callable[[str], Any],
    loose: bool = False,
) -> List[Dict[str, str]]:
    results: List[Dict[str, str]] = []

    for index, (query, expected) in enumerate(cases, start=1):
        try:
            actual_raw = func(query)
            actual = cell_to_str(actual_raw)
            passed = compare_output(expected, actual, loose=loose)
            error = ""
        except Exception:
            actual = ""
            passed = False
            error = traceback.format_exc(limit=3)

        results.append(
            {
                "序号": index,
                "输入": query,
                "预期输出": expected,
                "实际输出": actual,
                "是否符合预期": "是" if passed else "否",
                "错误信息": error,
            }
        )

    return results


def autosize_columns(ws, min_width: int = 10, max_width: int = 90) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = cell_to_str(cell.value)
            if value:
                max_len = max(max_len, min(len(value), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_width, max_len + 2)


def write_results(results: List[Dict[str, str]], output_path: str, loose: bool) -> None:
    total = len(results)
    passed = sum(1 for r in results if r["是否符合预期"] == "是")
    failed = total - passed
    pass_rate = passed / total if total else 0

    wb = Workbook()
    ws = wb.active
    ws.title = "回归测试结果"

    headers = ["序号", "输入", "预期输出", "实际输出", "是否符合预期", "错误信息"]
    ws.append(headers)

    for r in results:
        ws.append([r[h] for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    pass_fill = PatternFill("solid", fgColor="E2F0D9")
    fail_fill = PatternFill("solid", fgColor="FCE4D6")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status_cell = row[4]
        status_cell.fill = pass_fill if status_cell.value == "是" else fail_fill
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["F"].width = 60

    summary = wb.create_sheet("汇总")
    summary_rows = [
        ["指标", "值"],
        ["总用例数", total],
        ["通过数", passed],
        ["失败数", failed],
        ["通过率", f"{pass_rate:.2%}"],
        ["比较模式", "宽松比较" if loose else "严格比较"],
        ["核心口径", "模型系列版本并入 model-name；参数规模、量化、任务/模态等进入 attribute-tag"],
        ["示例", "qwen3-30B -> model-name: qwen3; attribute-tag: 30b"],
    ]
    for row in summary_rows:
        summary.append(row)

    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in summary.iter_rows(min_row=2, max_row=summary.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 100

    failed_ws = wb.create_sheet("失败用例")
    failed_ws.append(headers)
    for r in results:
        if r["是否符合预期"] != "是":
            failed_ws.append([r[h] for h in headers])

    for cell in failed_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in failed_ws.iter_rows(min_row=2, max_row=failed_ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    failed_ws.freeze_panes = "A2"
    autosize_columns(failed_ws)
    failed_ws.column_dimensions["C"].width = 80
    failed_ws.column_dimensions["D"].width = 80
    failed_ws.column_dimensions["F"].width = 60

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="模型识别 Agent 回归测试脚本 v2")
    parser.add_argument("--cases", default=DEFAULT_CASES, help=f"测试用例 xlsx 路径，默认：{DEFAULT_CASES}")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help=f"回归结果 xlsx 路径，默认：{DEFAULT_OUTPUT}")
    parser.add_argument("--module", required=True, help="包含 from_query_get_result 的 Python 模块名，例如 agent")
    parser.add_argument("--function", default="from_query_get_result", help="被测函数名，默认：from_query_get_result")
    parser.add_argument("--sheet", default=None, help="测试用例 Sheet 名；默认读取第一个 Sheet")
    parser.add_argument("--loose", action="store_true", help="启用宽松比较，兼容标点、空白、framework/framwork 等格式差异")
    parser.add_argument("--fail-on-error", action="store_true", help="存在失败用例时返回非 0 退出码，适合 CI 使用")
    args = parser.parse_args()

    func = load_function(args.module, args.function)
    cases = read_cases(args.cases, sheet_name=args.sheet)
    results = run_cases(cases, func, loose=args.loose)
    write_results(results, args.output, loose=args.loose)

    total = len(results)
    passed = sum(1 for r in results if r["是否符合预期"] == "是")
    failed = total - passed

    print("=" * 80)
    print("模型识别 Agent 回归测试完成")
    print(f"测试用例：{args.cases}")
    print(f"输出结果：{args.output}")
    print(f"总用例数：{total}")
    print(f"通过数：{passed}")
    print(f"失败数：{failed}")
    print(f"通过率：{(passed / total if total else 0):.2%}")
    print(f"比较模式：{'宽松比较' if args.loose else '严格比较'}")
    print("=" * 80)

    if failed:
        print("前 5 条失败用例：")
        for r in [x for x in results if x["是否符合预期"] != "是"][:5]:
            print("-" * 80)
            print(f"序号：{r['序号']}")
            print(f"输入：{r['输入']}")
            print(f"预期：{r['预期输出']}")
            print(f"实际：{r['实际输出']}")
            if r["错误信息"]:
                print(f"错误：{r['错误信息']}")

    if args.fail_on_error and failed:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
