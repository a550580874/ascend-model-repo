#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
模型识别 Agent 回归测试脚本

用途：
1. 读取测试用例 xlsx，默认读取第一个 sheet 的「输入」「预期输出」两列。
2. 调用你提供的 from_query_get_result(query)。
3. 生成新的 xlsx，包含「输入」「预期输出」「实际输出」「是否符合预期」四列。

示例：
    python run_model_recognition_regression.py \
      --cases 模型识别Agent_全量回归测试用例.xlsx \
      --module your_agent_module \
      --function from_query_get_result \
      --output 模型识别Agent_回归测试结果.xlsx

如果你的函数就在当前目录 agent.py 里：
    python run_model_recognition_regression.py --module agent
"""

from __future__ import annotations

import argparse
import importlib
import re
import sys
from pathlib import Path
from typing import Callable, Optional, Tuple

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "缺少依赖 openpyxl，请先安装：pip install openpyxl"
    ) from exc


def normalize_text(value: object, *, loose: bool = False) -> str:
    """用于比较的规整逻辑：默认 lower + 空白规整。"""
    if value is None:
        text = ""
    else:
        text = str(value)

    text = text.strip().lower()

    # 常见全角/中文标点归一化，降低格式性误差。
    replacements = {
        "；": ";",
        "，": ",",
        "：": ":",
        "（": "(",
        "）": ")",
        "【": "[",
        "】": "]",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)

    # 规整多余空白。
    text = re.sub(r"\s+", " ", text)
    # 规整分号两侧空格。
    text = re.sub(r"\s*;\s*", "; ", text).strip()
    # 去掉末尾多余分号。
    text = re.sub(r";\s*$", "", text)

    if loose:
        # 可选：兼容 framework / framwork 拼写差异，以及中英标签轻微差异。
        text = text.replace("framework:", "framwork:")
        text = text.replace("适配框架:", "framwork:")
        text = text.replace("模型名:", "model-name:")
        text = text.replace("属性标签:", "attribute-tag:")
        text = text.replace("硬件:", "hardware:")
        text = text.replace("[查询", "[query ")
        text = re.sub(r"\[query\s*(\d+)\]", r"[query \1]", text)
        text = text.replace("[query ", "[query ")
        text = re.sub(r"\s+", " ", text).strip()

    return text


def import_target_function(module_name: str, function_name: str) -> Callable[[str], object]:
    try:
        module = importlib.import_module(module_name)
    except Exception as exc:
        raise SystemExit(f"导入模块失败: {module_name}\n{exc}") from exc

    func = getattr(module, function_name, None)
    if func is None or not callable(func):
        raise SystemExit(f"模块 {module_name} 中找不到可调用函数: {function_name}")
    return func


def find_header_indexes(header_row) -> Tuple[int, int]:
    """返回 输入列、预期输出列的 1-based column index。"""
    normalized = [normalize_text(c.value) for c in header_row]

    input_candidates = {"输入", "input", "query", "用户输入"}
    expected_candidates = {"预期输出", "expected", "expected output", "expected_output"}

    input_idx: Optional[int] = None
    expected_idx: Optional[int] = None

    for i, name in enumerate(normalized, start=1):
        if name in input_candidates and input_idx is None:
            input_idx = i
        if name in expected_candidates and expected_idx is None:
            expected_idx = i

    if input_idx is None or expected_idx is None:
        raise SystemExit(
            "测试用例表头必须包含「输入」「预期输出」两列，"
            "或英文 input / expected_output。"
        )
    return input_idx, expected_idx


def autosize_columns(ws, max_width: int = 90) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 12
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def write_result_xlsx(results, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "回归测试结果"

    headers = ["输入", "预期输出", "实际输出", "是否符合预期"]
    ws.append(headers)

    for row in results:
        ws.append([row["input"], row["expected"], row["actual"], row["passed"]])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    pass_fill = PatternFill("solid", fgColor="E2F0D9")
    fail_fill = PatternFill("solid", fgColor="FCE4D6")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        row[3].fill = pass_fill if row[3].value == "是" else fail_fill

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    ws.column_dimensions["A"].width = min(ws.column_dimensions["A"].width, 50)
    ws.column_dimensions["B"].width = min(ws.column_dimensions["B"].width, 100)
    ws.column_dimensions["C"].width = min(ws.column_dimensions["C"].width, 100)
    ws.column_dimensions["D"].width = 14
    ws.auto_filter.ref = ws.dimensions

    # 汇总 sheet
    summary = wb.create_sheet("汇总")
    total = len(results)
    passed = sum(1 for r in results if r["passed"] == "是")
    failed = total - passed
    summary_rows = [
        ["指标", "值"],
        ["总用例数", total],
        ["通过数", passed],
        ["失败数", failed],
        ["通过率", f"{passed / total:.2%}" if total else "0.00%"],
    ]
    for row in summary_rows:
        summary.append(row)
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    autosize_columns(summary, max_width=30)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def run_regression(
    cases_path: Path,
    output_path: Path,
    target_func: Callable[[str], object],
    sheet_name: Optional[str],
    loose_compare: bool,
) -> None:
    wb = load_workbook(cases_path)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    input_idx, expected_idx = find_header_indexes(ws[1])
    results = []

    for row_num in range(2, ws.max_row + 1):
        query = ws.cell(row_num, input_idx).value
        expected = ws.cell(row_num, expected_idx).value
        if query is None or str(query).strip() == "":
            continue

        try:
            actual_obj = target_func(str(query))
            actual = "" if actual_obj is None else str(actual_obj)
        except Exception as exc:
            actual = f"__exception__: {type(exc).__name__}: {exc}"

        passed = normalize_text(actual, loose=loose_compare) == normalize_text(expected, loose=loose_compare)
        results.append(
            {
                "input": str(query),
                "expected": "" if expected is None else str(expected),
                "actual": actual,
                "passed": "是" if passed else "否",
            }
        )

    write_result_xlsx(results, output_path)

    total = len(results)
    passed_count = sum(1 for r in results if r["passed"] == "是")
    failed_count = total - passed_count
    print(f"总用例数: {total}")
    print(f"通过数: {passed_count}")
    print(f"失败数: {failed_count}")
    print(f"通过率: {passed_count / total:.2%}" if total else "通过率: 0.00%")
    print(f"结果文件: {output_path}")

    if failed_count:
        print("\n失败样例预览：")
        shown = 0
        for r in results:
            if r["passed"] == "否":
                print("- 输入:", r["input"])
                print("  预期:", r["expected"])
                print("  实际:", r["actual"])
                shown += 1
                if shown >= 5:
                    break


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="模型识别 Agent 回归测试")
    parser.add_argument("--cases", default="模型识别Agent_全量回归测试用例.xlsx", help="测试用例 xlsx 路径")
    parser.add_argument("--output", default="模型识别Agent_回归测试结果.xlsx", help="输出结果 xlsx 路径")
    parser.add_argument("--module", required=True, help="包含 from_query_get_result 的 Python 模块名，例如 agent")
    parser.add_argument("--function", default="from_query_get_result", help="目标函数名")
    parser.add_argument("--sheet", default=None, help="测试用例 sheet 名；默认第一个 sheet")
    parser.add_argument("--loose", action="store_true", help="启用宽松比较：兼容 framework/framwork 和中英标签差异")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    target_func = import_target_function(args.module, args.function)
    run_regression(
        cases_path=Path(args.cases),
        output_path=Path(args.output),
        target_func=target_func,
        sheet_name=args.sheet,
        loose_compare=args.loose,
    )


if __name__ == "__main__":
    main(sys.argv[1:])
